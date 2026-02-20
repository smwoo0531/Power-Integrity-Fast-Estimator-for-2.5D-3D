import re
from dataclasses import dataclass
from typing import Any, List, Tuple

import numpy as np
from openpyxl import load_workbook


_CELL_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _col_to_idx(col: str) -> int:
    col = col.upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _cell_to_rc(cell: str) -> Tuple[int, int]:
    m = _CELL_RE.match(cell.strip())
    if not m:
        raise ValueError(f"Invalid cell reference: {cell}")
    col, row = m.group(1), int(m.group(2))
    return row, _col_to_idx(col)


def _parse_range(rng: str) -> Tuple[int, int, int, int]:
    parts = rng.split(":")
    if len(parts) == 1:
        r1, c1 = _cell_to_rc(parts[0])
        return r1, c1, r1, c1
    if len(parts) != 2:
        raise ValueError(f"Invalid range: {rng}")
    r1, c1 = _cell_to_rc(parts[0])
    r2, c2 = _cell_to_rc(parts[1])
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)


@dataclass
class ExcelBook:
    path: str
    data_only: bool = True

    def __post_init__(self) -> None:
        self._wb = load_workbook(self.path, data_only=self.data_only)

    def _get_sheet(self, sheet: str):
        if sheet not in self._wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in {self.path}")
        return self._wb[sheet]

    def read_range(self, sheet: str, rng: str) -> np.ndarray:
        ws = self._get_sheet(sheet)
        r1, c1, r2, c2 = _parse_range(rng)
        rows = []
        for r in range(r1, r2 + 1):
            row_vals = []
            for c in range(c1, c2 + 1):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    val = np.nan
                row_vals.append(val)
            rows.append(row_vals)
        arr = np.array(rows, dtype=float)
        if arr.shape[0] == 1 and arr.shape[1] == 1:
            return arr.reshape(())
        if 1 in arr.shape:
            return arr.flatten()
        return arr

    def read_scalar(self, sheet: str, rng: str) -> float:
        val = self.read_range(sheet, rng)
        if isinstance(val, np.ndarray):
            return float(val.reshape(-1)[0])
        return float(val)

    def read_vector(self, sheet: str, rng: str) -> np.ndarray:
        arr = self.read_range(sheet, rng)
        if isinstance(arr, np.ndarray):
            return arr.astype(float)
        return np.array([arr], dtype=float)
